"""Unit tests for the export service layer.

Tests cover CSV and XLSX exports for customers, inventory items,
service orders, and invoices.  CSV tests parse the returned string
and verify headers and row counts.  XLSX tests verify the returned
BytesIO buffer can be loaded by openpyxl.
"""

import csv
import io
from datetime import date
from decimal import Decimal

import pytest

from app.extensions import db
from app.models.customer import Customer
from app.models.inventory import InventoryItem
from app.models.invoice import Invoice
from app.models.service_order import ServiceOrder
from app.services import export_service
from tests.factories import (
    CustomerFactory,
    InventoryItemFactory,
    InvoiceFactory,
    ServiceOrderFactory,
)

pytestmark = pytest.mark.unit


def _set_session(db_session):
    """Configure all factories to use the given session."""
    CustomerFactory._meta.sqlalchemy_session = db_session
    InventoryItemFactory._meta.sqlalchemy_session = db_session
    InvoiceFactory._meta.sqlalchemy_session = db_session
    ServiceOrderFactory._meta.sqlalchemy_session = db_session


def _parse_csv(csv_string):
    """Parse a CSV string (with possible BOM) into a list of rows."""
    # Strip UTF-8 BOM if present
    clean = csv_string.lstrip("\ufeff")
    reader = csv.reader(io.StringIO(clean))
    return list(reader)


# =========================================================================
# CSV — Customers
# =========================================================================


class TestExportCustomersCsv:
    """Tests for export_customers_csv()."""

    def test_export_customers_csv(self, app, db_session):
        """Exports customers to CSV with correct headers and data rows."""
        _set_session(db_session)
        CustomerFactory(first_name="Alice", last_name="Diver")
        CustomerFactory(first_name="Bob", last_name="Swimmer")

        result = export_service.export_customers_csv()

        rows = _parse_csv(result)
        headers = rows[0]
        assert "ID" in headers
        assert "First Name" in headers
        assert "Last Name" in headers
        assert "Email" in headers
        # 1 header + 2 data rows
        assert len(rows) == 3

    def test_export_customers_csv_empty(self, app, db_session):
        """Exports only headers when no customers exist."""
        result = export_service.export_customers_csv()

        rows = _parse_csv(result)
        # Only the header row
        assert len(rows) == 1
        assert "ID" in rows[0]


# =========================================================================
# CSV — Inventory
# =========================================================================


class TestExportInventoryCsv:
    """Tests for export_inventory_csv()."""

    def test_export_inventory_csv(self, app, db_session):
        """Exports inventory items to CSV with correct headers and data."""
        _set_session(db_session)
        InventoryItemFactory(name="O-Ring", sku="SKU-EXP-01")
        InventoryItemFactory(name="Neck Seal", sku="SKU-EXP-02")

        result = export_service.export_inventory_csv()

        rows = _parse_csv(result)
        headers = rows[0]
        assert "ID" in headers
        assert "SKU" in headers
        assert "Name" in headers
        assert "Category" in headers
        # 1 header + 2 data rows
        assert len(rows) == 3


# =========================================================================
# CSV — Orders
# =========================================================================


class TestExportOrdersCsv:
    """Tests for export_orders_csv()."""

    def test_export_orders_csv(self, app, db_session):
        """Exports service orders to CSV with correct headers and data."""
        _set_session(db_session)
        ServiceOrderFactory(order_number="SO-2026-50001")
        ServiceOrderFactory(order_number="SO-2026-50002")

        result = export_service.export_orders_csv()

        rows = _parse_csv(result)
        headers = rows[0]
        assert "ID" in headers
        assert "Order Number" in headers
        assert "Status" in headers
        # 1 header + 2 data rows
        assert len(rows) == 3


# =========================================================================
# CSV — Invoices
# =========================================================================


class TestExportInvoicesCsv:
    """Tests for export_invoices_csv()."""

    def test_export_invoices_csv(self, app, db_session):
        """Exports invoices to CSV with correct headers and data."""
        _set_session(db_session)
        InvoiceFactory(invoice_number="INV-2026-50001")
        InvoiceFactory(invoice_number="INV-2026-50002")

        result = export_service.export_invoices_csv()

        rows = _parse_csv(result)
        headers = rows[0]
        assert "ID" in headers
        assert "Invoice Number" in headers
        assert "Status" in headers
        assert "Total" in headers
        # 1 header + 2 data rows
        assert len(rows) == 3


# =========================================================================
# XLSX — Customers
# =========================================================================


class TestExportCustomersXlsx:
    """Tests for export_customers_xlsx()."""

    def test_export_customers_xlsx(self, app, db_session):
        """Exports customers to XLSX with correct headers."""
        _set_session(db_session)
        CustomerFactory(first_name="Excel", last_name="Test")

        result = export_service.export_customers_xlsx()

        assert isinstance(result, io.BytesIO)

        from openpyxl import load_workbook
        wb = load_workbook(result)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "ID" in headers
        assert "First Name" in headers
        assert "Last Name" in headers
        # At least 1 data row
        assert ws.max_row >= 2


# =========================================================================
# XLSX — Inventory
# =========================================================================


class TestExportInventoryXlsx:
    """Tests for export_inventory_xlsx()."""

    def test_export_inventory_xlsx(self, app, db_session):
        """Exports inventory items to XLSX with correct headers."""
        _set_session(db_session)
        InventoryItemFactory(name="Zipper", sku="SKU-XLS-01")

        result = export_service.export_inventory_xlsx()

        assert isinstance(result, io.BytesIO)

        from openpyxl import load_workbook
        wb = load_workbook(result)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "ID" in headers
        assert "SKU" in headers
        assert "Name" in headers
        assert ws.max_row >= 2


# =========================================================================
# XLSX — Orders
# =========================================================================


class TestExportOrdersXlsx:
    """Tests for export_orders_xlsx()."""

    def test_export_orders_xlsx(self, app, db_session):
        """Exports service orders to XLSX with correct headers."""
        _set_session(db_session)
        ServiceOrderFactory(order_number="SO-2026-60001")

        result = export_service.export_orders_xlsx()

        assert isinstance(result, io.BytesIO)

        from openpyxl import load_workbook
        wb = load_workbook(result)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "ID" in headers
        assert "Order Number" in headers
        assert "Status" in headers
        assert ws.max_row >= 2


# =========================================================================
# XLSX — Invoices
# =========================================================================


class TestExportInvoicesXlsx:
    """Tests for export_invoices_xlsx()."""

    def test_export_invoices_xlsx(self, app, db_session):
        """Exports invoices to XLSX with correct headers."""
        _set_session(db_session)
        InvoiceFactory(invoice_number="INV-2026-60001")

        result = export_service.export_invoices_xlsx()

        assert isinstance(result, io.BytesIO)

        from openpyxl import load_workbook
        wb = load_workbook(result)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "ID" in headers
        assert "Invoice Number" in headers
        assert "Total" in headers
        assert ws.max_row >= 2
