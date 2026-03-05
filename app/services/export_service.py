"""Export service -- CSV and XLSX export generation for entity data.

Provides functions to export customers, inventory items, service orders,
and invoices in both CSV and XLSX formats.  CSV functions return a UTF-8
string (with BOM for Excel compatibility).  XLSX functions return an
``io.BytesIO`` buffer containing a complete workbook.

XLSX support requires the ``openpyxl`` package.  If openpyxl is not
installed, the XLSX functions will raise ``RuntimeError`` while the CSV
functions continue to work normally.
"""

import csv
import io
from datetime import date, datetime
from decimal import Decimal

from app.extensions import db
from app.models.customer import Customer
from app.models.inventory import InventoryItem
from app.models.invoice import Invoice
from app.models.service_order import ServiceOrder

# ---------------------------------------------------------------------------
# Optional openpyxl import -- XLSX exports degrade gracefully
# ---------------------------------------------------------------------------

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:  # pragma: no cover
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _format_value(val):
    """Format a value for export.

    * ``None`` -> empty string
    * ``date`` / ``datetime`` -> ISO-8601 string
    * ``Decimal`` -> plain string (no scientific notation)
    * ``bool`` -> "Yes" / "No"
    * Everything else -> ``str(val)``
    """
    if val is None:
        return ""
    if isinstance(val, (date, datetime)):
        return val.isoformat()
    if isinstance(val, Decimal):
        return str(val)
    if isinstance(val, bool):
        return "Yes" if val else "No"
    return str(val)


def _require_openpyxl():
    """Raise ``RuntimeError`` if openpyxl is not available."""
    if not HAS_OPENPYXL:
        raise RuntimeError(
            "XLSX export requires the 'openpyxl' package.  "
            "Install it with: pip install openpyxl"
        )


def _create_xlsx(headers, rows):
    """Build an XLSX workbook from *headers* and *rows*.

    Returns an ``io.BytesIO`` buffer with the workbook saved to it and
    the stream position reset to the beginning.

    Parameters
    ----------
    headers : list[str]
        Column header labels.
    rows : list[list]
        Data rows, each the same length as *headers*.
    """
    _require_openpyxl()

    wb = Workbook()
    ws = wb.active

    # -- Write header row with bold font --
    bold = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    # -- Write data rows --
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # -- Auto-size columns (approximate) --
    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(headers[col_idx - 1]))
        for row_idx in range(2, len(rows) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        # Cap at a reasonable maximum and add a small padding
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # -- Save to buffer --
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ===================================================================
# CSV exports
# ===================================================================


def export_customers_csv(query=None):
    """Export customers to CSV format.

    Args:
        query: Optional pre-filtered SQLAlchemy query.  If ``None``,
               exports all non-deleted customers ordered by last name.

    Returns:
        A string containing CSV data with a UTF-8 BOM.
    """
    if query is None:
        query = Customer.not_deleted().order_by(Customer.last_name)

    customers = query.all()

    output = io.StringIO()
    output.write("\ufeff")  # UTF-8 BOM for Excel compatibility

    writer = csv.writer(output)
    headers = [
        "ID", "Type", "First Name", "Last Name", "Business Name",
        "Contact Person", "Email", "Phone", "Address", "City",
        "State", "Postal Code", "Country", "Preferred Contact",
        "Tax Exempt", "Notes", "Created",
    ]
    writer.writerow(headers)

    for c in customers:
        writer.writerow([
            c.id, c.customer_type, c.first_name, c.last_name,
            c.business_name, c.contact_person, c.email,
            c.phone_primary, c.address_line1, c.city,
            c.state_province, c.postal_code, c.country,
            c.preferred_contact, _format_value(c.tax_exempt),
            c.notes, _format_value(c.created_at),
        ])

    return output.getvalue()


def export_inventory_csv(query=None):
    """Export inventory items to CSV format.

    Args:
        query: Optional pre-filtered SQLAlchemy query.  If ``None``,
               exports all non-deleted items ordered by name.

    Returns:
        A string containing CSV data with a UTF-8 BOM.
    """
    if query is None:
        query = InventoryItem.query.filter_by(is_deleted=False).order_by(
            InventoryItem.name
        )

    items = query.all()
    output = io.StringIO()
    output.write("\ufeff")

    writer = csv.writer(output)
    headers = [
        "ID", "SKU", "Name", "Category", "Subcategory",
        "Manufacturer", "Purchase Cost", "Resale Price",
        "Quantity", "Reorder Level", "Unit", "Location",
        "Active", "Notes", "Created",
    ]
    writer.writerow(headers)

    for item in items:
        writer.writerow([
            item.id, item.sku, item.name, item.category,
            item.subcategory, item.manufacturer,
            _format_value(item.purchase_cost), _format_value(item.resale_price),
            item.quantity_in_stock, item.reorder_level,
            item.unit_of_measure, item.storage_location,
            _format_value(item.is_active), item.notes,
            _format_value(item.created_at),
        ])

    return output.getvalue()


def export_orders_csv(query=None):
    """Export service orders to CSV format.

    Args:
        query: Optional pre-filtered SQLAlchemy query.  If ``None``,
               exports all non-deleted orders ordered by date received
               (newest first).

    Returns:
        A string containing CSV data with a UTF-8 BOM.
    """
    if query is None:
        query = ServiceOrder.not_deleted().order_by(
            ServiceOrder.date_received.desc()
        )

    orders = query.all()
    output = io.StringIO()
    output.write("\ufeff")

    writer = csv.writer(output)
    headers = [
        "ID", "Order Number", "Customer", "Status", "Priority",
        "Assigned Tech", "Date Received", "Date Promised",
        "Date Completed", "Estimated Total", "Description", "Created",
    ]
    writer.writerow(headers)

    for o in orders:
        writer.writerow([
            o.id, o.order_number,
            o.customer.display_name if o.customer else "",
            o.status, o.priority,
            o.assigned_tech.username if o.assigned_tech else "",
            _format_value(o.date_received), _format_value(o.date_promised),
            _format_value(o.date_completed), _format_value(o.estimated_total),
            o.description, _format_value(o.created_at),
        ])

    return output.getvalue()


def export_invoices_csv(query=None):
    """Export invoices to CSV format.

    Args:
        query: Optional pre-filtered SQLAlchemy query.  If ``None``,
               exports all invoices ordered by issue date (newest first).

    Returns:
        A string containing CSV data with a UTF-8 BOM.
    """
    if query is None:
        query = Invoice.query.order_by(Invoice.issue_date.desc())

    invoices = query.all()
    output = io.StringIO()
    output.write("\ufeff")

    writer = csv.writer(output)
    headers = [
        "ID", "Invoice Number", "Customer", "Status",
        "Issue Date", "Due Date", "Subtotal", "Tax",
        "Discount", "Total", "Amount Paid", "Balance Due",
        "Notes", "Created",
    ]
    writer.writerow(headers)

    for inv in invoices:
        writer.writerow([
            inv.id, inv.invoice_number,
            inv.customer.display_name if inv.customer else "",
            inv.status, _format_value(inv.issue_date),
            _format_value(inv.due_date), _format_value(inv.subtotal),
            _format_value(inv.tax_amount), _format_value(inv.discount_amount),
            _format_value(inv.total), _format_value(inv.amount_paid),
            _format_value(inv.balance_due), inv.notes,
            _format_value(inv.created_at),
        ])

    return output.getvalue()


# ===================================================================
# XLSX exports
# ===================================================================


def export_customers_xlsx(query=None):
    """Export customers to XLSX format.

    Args:
        query: Optional pre-filtered SQLAlchemy query.  If ``None``,
               exports all non-deleted customers ordered by last name.

    Returns:
        An ``io.BytesIO`` buffer containing the XLSX workbook.

    Raises:
        RuntimeError: If openpyxl is not installed.
    """
    if query is None:
        query = Customer.not_deleted().order_by(Customer.last_name)

    customers = query.all()

    headers = [
        "ID", "Type", "First Name", "Last Name", "Business Name",
        "Contact Person", "Email", "Phone", "Address", "City",
        "State", "Postal Code", "Country", "Preferred Contact",
        "Tax Exempt", "Notes", "Created",
    ]

    rows = []
    for c in customers:
        rows.append([
            c.id, c.customer_type, c.first_name, c.last_name,
            c.business_name, c.contact_person, c.email,
            c.phone_primary, c.address_line1, c.city,
            c.state_province, c.postal_code, c.country,
            c.preferred_contact, _format_value(c.tax_exempt),
            c.notes, _format_value(c.created_at),
        ])

    return _create_xlsx(headers, rows)


def export_inventory_xlsx(query=None):
    """Export inventory items to XLSX format.

    Args:
        query: Optional pre-filtered SQLAlchemy query.  If ``None``,
               exports all non-deleted items ordered by name.

    Returns:
        An ``io.BytesIO`` buffer containing the XLSX workbook.

    Raises:
        RuntimeError: If openpyxl is not installed.
    """
    if query is None:
        query = InventoryItem.query.filter_by(is_deleted=False).order_by(
            InventoryItem.name
        )

    items = query.all()

    headers = [
        "ID", "SKU", "Name", "Category", "Subcategory",
        "Manufacturer", "Purchase Cost", "Resale Price",
        "Quantity", "Reorder Level", "Unit", "Location",
        "Active", "Notes", "Created",
    ]

    rows = []
    for item in items:
        rows.append([
            item.id, item.sku, item.name, item.category,
            item.subcategory, item.manufacturer,
            _format_value(item.purchase_cost), _format_value(item.resale_price),
            item.quantity_in_stock, item.reorder_level,
            item.unit_of_measure, item.storage_location,
            _format_value(item.is_active), item.notes,
            _format_value(item.created_at),
        ])

    return _create_xlsx(headers, rows)


def export_orders_xlsx(query=None):
    """Export service orders to XLSX format.

    Args:
        query: Optional pre-filtered SQLAlchemy query.  If ``None``,
               exports all non-deleted orders ordered by date received
               (newest first).

    Returns:
        An ``io.BytesIO`` buffer containing the XLSX workbook.

    Raises:
        RuntimeError: If openpyxl is not installed.
    """
    if query is None:
        query = ServiceOrder.not_deleted().order_by(
            ServiceOrder.date_received.desc()
        )

    orders = query.all()

    headers = [
        "ID", "Order Number", "Customer", "Status", "Priority",
        "Assigned Tech", "Date Received", "Date Promised",
        "Date Completed", "Estimated Total", "Description", "Created",
    ]

    rows = []
    for o in orders:
        rows.append([
            o.id, o.order_number,
            o.customer.display_name if o.customer else "",
            o.status, o.priority,
            o.assigned_tech.username if o.assigned_tech else "",
            _format_value(o.date_received), _format_value(o.date_promised),
            _format_value(o.date_completed), _format_value(o.estimated_total),
            o.description, _format_value(o.created_at),
        ])

    return _create_xlsx(headers, rows)


def export_invoices_xlsx(query=None):
    """Export invoices to XLSX format.

    Args:
        query: Optional pre-filtered SQLAlchemy query.  If ``None``,
               exports all invoices ordered by issue date (newest first).

    Returns:
        An ``io.BytesIO`` buffer containing the XLSX workbook.

    Raises:
        RuntimeError: If openpyxl is not installed.
    """
    if query is None:
        query = Invoice.query.order_by(Invoice.issue_date.desc())

    invoices = query.all()

    headers = [
        "ID", "Invoice Number", "Customer", "Status",
        "Issue Date", "Due Date", "Subtotal", "Tax",
        "Discount", "Total", "Amount Paid", "Balance Due",
        "Notes", "Created",
    ]

    rows = []
    for inv in invoices:
        rows.append([
            inv.id, inv.invoice_number,
            inv.customer.display_name if inv.customer else "",
            inv.status, _format_value(inv.issue_date),
            _format_value(inv.due_date), _format_value(inv.subtotal),
            _format_value(inv.tax_amount), _format_value(inv.discount_amount),
            _format_value(inv.total), _format_value(inv.amount_paid),
            _format_value(inv.balance_due), inv.notes,
            _format_value(inv.created_at),
        ])

    return _create_xlsx(headers, rows)
