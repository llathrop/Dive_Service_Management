"""Import service — CSV/XLSX import for customers and inventory.

Supports both fixed column format (legacy) and interactive column mapping.
Import workflow: detect columns → map → validate → preview → commit.
"""

import csv
import io
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher

from app.extensions import db
from app.models.customer import Customer
from app.models.inventory import InventoryItem


# Expected column headers (case-insensitive, order matters for mapping)
CUSTOMER_COLUMNS = [
    "type", "first_name", "last_name", "business_name",
    "contact_person", "email", "phone", "address", "city",
    "state", "postal_code", "country", "preferred_contact",
    "tax_exempt", "notes",
]

INVENTORY_COLUMNS = [
    "sku", "name", "category", "subcategory",
    "manufacturer", "purchase_cost", "resale_price",
    "quantity", "reorder_level", "unit", "location",
    "notes",
]

# Target field definitions with labels and required flags
CUSTOMER_FIELDS = [
    {"key": "type", "label": "Type", "required": False},
    {"key": "first_name", "label": "First Name", "required": False},
    {"key": "last_name", "label": "Last Name", "required": False},
    {"key": "business_name", "label": "Business Name", "required": False},
    {"key": "contact_person", "label": "Contact Person", "required": False},
    {"key": "email", "label": "Email", "required": False},
    {"key": "phone", "label": "Phone", "required": False},
    {"key": "address", "label": "Address", "required": False},
    {"key": "city", "label": "City", "required": False},
    {"key": "state", "label": "State", "required": False},
    {"key": "postal_code", "label": "Postal Code", "required": False},
    {"key": "country", "label": "Country", "required": False},
    {"key": "preferred_contact", "label": "Preferred Contact", "required": False},
    {"key": "tax_exempt", "label": "Tax Exempt", "required": False},
    {"key": "notes", "label": "Notes", "required": False},
]

INVENTORY_FIELDS = [
    {"key": "sku", "label": "SKU", "required": False},
    {"key": "name", "label": "Name", "required": True},
    {"key": "category", "label": "Category", "required": True},
    {"key": "subcategory", "label": "Subcategory", "required": False},
    {"key": "manufacturer", "label": "Manufacturer", "required": False},
    {"key": "purchase_cost", "label": "Purchase Cost", "required": False},
    {"key": "resale_price", "label": "Resale Price", "required": False},
    {"key": "quantity", "label": "Quantity", "required": False},
    {"key": "reorder_level", "label": "Reorder Level", "required": False},
    {"key": "unit", "label": "Unit", "required": False},
    {"key": "location", "label": "Location", "required": False},
    {"key": "notes", "label": "Notes", "required": False},
]


# ---------------------------------------------------------------------------
# Column mapping wizard functions
# ---------------------------------------------------------------------------

def detect_columns(file_content, file_type="csv"):
    """Read the first row of a CSV or XLSX file and return column names.

    Args:
        file_content: For CSV, a string. For XLSX, raw bytes.
        file_type: "csv" or "xlsx".

    Returns:
        list of column name strings, or empty list on error.
    """
    if file_type == "xlsx":
        return _detect_columns_xlsx(file_content)
    return _detect_columns_csv(file_content)


def _detect_columns_csv(file_content):
    """Detect columns from CSV string content."""
    if isinstance(file_content, bytes):
        file_content = file_content.decode("utf-8-sig")
    if file_content.startswith("\ufeff"):
        file_content = file_content[1:]

    reader = csv.reader(io.StringIO(file_content))
    try:
        headers = next(reader)
        return [h.strip() for h in headers if h.strip()]
    except StopIteration:
        return []


def _detect_columns_xlsx(file_content):
    """Detect columns from XLSX bytes."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
        ws = wb.active
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(cell).strip() if cell is not None else "" for cell in row]
        wb.close()
        return [h for h in headers if h]
    except Exception:
        return []


def get_target_fields(entity_type):
    """Return list of valid target field dicts for the given entity type.

    Each dict has keys: "key", "label", "required".

    Args:
        entity_type: "customers" or "inventory".

    Returns:
        list of field definition dicts.
    """
    if entity_type == "customers":
        return list(CUSTOMER_FIELDS)
    elif entity_type == "inventory":
        return list(INVENTORY_FIELDS)
    return []


def auto_detect_mapping(source_columns, entity_type):
    """Guess column mappings by comparing source column names to target fields.

    Uses fuzzy string matching (SequenceMatcher) to find the best match.

    Args:
        source_columns: list of source column name strings.
        entity_type: "customers" or "inventory".

    Returns:
        dict mapping source column name -> target field key (or None for no match).
    """
    target_fields = get_target_fields(entity_type)
    mapping = {}

    # Build lookup: normalize target labels and keys for matching
    targets = []
    for field in target_fields:
        targets.append({
            "key": field["key"],
            "label": field["label"],
            "normalized": _normalize(field["label"]),
            "key_normalized": _normalize(field["key"]),
        })

    used_targets = set()

    for source_col in source_columns:
        source_norm = _normalize(source_col)
        best_key = None
        best_score = 0.0

        for target in targets:
            if target["key"] in used_targets:
                continue

            # Try matching against label and key
            score_label = SequenceMatcher(
                None, source_norm, target["normalized"]
            ).ratio()
            score_key = SequenceMatcher(
                None, source_norm, target["key_normalized"]
            ).ratio()
            score = max(score_label, score_key)

            # Exact match bonus
            if source_norm == target["normalized"] or source_norm == target["key_normalized"]:
                score = 1.0

            if score > best_score:
                best_score = score
                best_key = target["key"]

        # Only accept matches above threshold
        if best_score >= 0.6 and best_key:
            mapping[source_col] = best_key
            used_targets.add(best_key)
        else:
            mapping[source_col] = None  # Skip/ignore

    return mapping


def _normalize(s):
    """Normalize a string for fuzzy matching: lowercase, remove underscores/spaces."""
    return s.lower().replace("_", " ").replace("-", " ").strip()


def _read_rows(file_content, file_type="csv"):
    """Read all data rows from CSV or XLSX content.

    Returns:
        tuple of (headers_list, rows_list_of_lists).
    """
    if file_type == "xlsx":
        return _read_rows_xlsx(file_content)
    return _read_rows_csv(file_content)


def _read_rows_csv(file_content):
    """Read rows from CSV string/bytes."""
    if isinstance(file_content, bytes):
        file_content = file_content.decode("utf-8-sig")
    if file_content.startswith("\ufeff"):
        file_content = file_content[1:]

    reader = csv.reader(io.StringIO(file_content))
    try:
        headers = next(reader)
        headers = [h.strip() for h in headers]
    except StopIteration:
        return [], []

    rows = []
    for row in reader:
        # Pad short rows, trim long rows
        padded = row + [""] * max(0, len(headers) - len(row))
        rows.append(padded[:len(headers)])
    return headers, rows


def _read_rows_xlsx(file_content):
    """Read rows from XLSX bytes."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not all_rows:
            return [], []

        headers = [str(cell).strip() if cell is not None else "" for cell in all_rows[0]]
        rows = []
        for raw_row in all_rows[1:]:
            row = [str(cell).strip() if cell is not None else "" for cell in raw_row]
            # Pad short rows
            padded = row + [""] * max(0, len(headers) - len(row))
            rows.append(padded[:len(headers)])
        return headers, rows
    except Exception:
        return [], []


def map_and_validate(file_content, column_mapping, entity_type, file_type="csv"):
    """Apply column mapping to file data and validate each row.

    Args:
        file_content: CSV string or XLSX bytes.
        column_mapping: dict mapping source column name -> target field key (or None to skip).
        entity_type: "customers" or "inventory".
        file_type: "csv" or "xlsx".

    Returns:
        dict with keys:
          - "headers": list of mapped target field keys (excluding skipped)
          - "rows": list of dicts (target_key -> value)
          - "errors": list of {"row": N, "message": str}
          - "preview": first 10 mapped rows as dicts
          - "total": total row count
          - "valid": count of rows without errors
    """
    headers, raw_rows = _read_rows(file_content, file_type)

    if not headers:
        return {
            "headers": [], "rows": [], "errors": [{"row": 0, "message": "Empty or invalid file"}],
            "preview": [], "total": 0, "valid": 0,
        }

    # Build index mapping: source column index -> target field key
    col_index_map = {}
    for i, header in enumerate(headers):
        target_key = column_mapping.get(header)
        if target_key:  # None or empty means skip
            col_index_map[i] = target_key

    # Map rows
    mapped_rows = []
    for raw_row in raw_rows:
        mapped = {}
        for col_idx, target_key in col_index_map.items():
            if col_idx < len(raw_row):
                val = raw_row[col_idx].strip() if raw_row[col_idx] else ""
                mapped[target_key] = val
            else:
                mapped[target_key] = ""
        mapped_rows.append(mapped)

    # Validate
    errors = []
    if entity_type == "customers":
        errors = _validate_mapped_customer_rows(mapped_rows)
    elif entity_type == "inventory":
        errors = _validate_mapped_inventory_rows(mapped_rows)

    # Determine which rows have errors
    error_row_nums = {e["row"] for e in errors}
    valid_count = sum(1 for i in range(len(mapped_rows)) if (i + 2) not in error_row_nums)

    target_headers = list(col_index_map.values())

    return {
        "headers": target_headers,
        "rows": mapped_rows,
        "errors": errors,
        "preview": mapped_rows[:10],
        "total": len(mapped_rows),
        "valid": valid_count,
    }


def _validate_mapped_customer_rows(rows):
    """Validate customer rows using mapped field keys."""
    errors = []
    for i, row in enumerate(rows, start=2):
        ctype = row.get("type", "").strip() or "individual"
        if ctype == "individual":
            if not row.get("first_name", "").strip() and not row.get("last_name", "").strip():
                errors.append({"row": i, "message": "Individual customers require first or last name"})
        elif ctype == "business":
            if not row.get("business_name", "").strip():
                errors.append({"row": i, "message": "Business customers require a business name"})
    return errors


def _validate_mapped_inventory_rows(rows):
    """Validate inventory rows using mapped field keys."""
    errors = []
    for i, row in enumerate(rows, start=2):
        if not row.get("name", "").strip():
            errors.append({"row": i, "message": "Name is required"})
        if not row.get("category", "").strip():
            errors.append({"row": i, "message": "Category is required"})
        cost = row.get("purchase_cost", "").strip()
        if cost:
            try:
                Decimal(cost)
            except InvalidOperation:
                errors.append({"row": i, "message": f"Invalid purchase cost: {cost}"})
        price = row.get("resale_price", "").strip()
        if price:
            try:
                Decimal(price)
            except InvalidOperation:
                errors.append({"row": i, "message": f"Invalid resale price: {price}"})
    return errors


def execute_mapped_import(file_content, column_mapping, entity_type, file_type="csv"):
    """Apply column mapping and import records into the database.

    Args:
        file_content: CSV string or XLSX bytes.
        column_mapping: dict mapping source column name -> target field key (or None).
        entity_type: "customers" or "inventory".
        file_type: "csv" or "xlsx".

    Returns:
        dict: {"imported": N, "skipped": N, "errors": [...]}
    """
    result = map_and_validate(file_content, column_mapping, entity_type, file_type)

    if entity_type == "customers":
        return _import_mapped_customers(result["rows"])
    elif entity_type == "inventory":
        return _import_mapped_inventory(result["rows"])

    return {"imported": 0, "skipped": 0, "errors": [{"row": 0, "message": f"Unknown entity type: {entity_type}"}]}


def _import_mapped_customers(rows):
    """Import customer rows using mapped field keys."""
    imported = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        try:
            email = row.get("email", "").strip()
            if email:
                existing = Customer.query.filter_by(email=email).first()
                if existing:
                    skipped += 1
                    continue

            first_name = row.get("first_name", "").strip()
            last_name = row.get("last_name", "").strip()
            ctype = row.get("type", "").strip() or "individual"

            # Validate required fields
            if ctype == "individual" and not first_name and not last_name:
                errors.append({"row": i, "message": "Individual customers require first or last name"})
                continue
            if ctype == "business" and not row.get("business_name", "").strip():
                errors.append({"row": i, "message": "Business customers require a business name"})
                continue

            customer = Customer(
                customer_type=ctype,
                first_name=first_name,
                last_name=last_name,
                business_name=row.get("business_name", "").strip() or None,
                contact_person=row.get("contact_person", "").strip() or None,
                email=email or None,
                phone_primary=row.get("phone", "").strip() or None,
                address_line1=row.get("address", "").strip() or None,
                city=row.get("city", "").strip() or None,
                state_province=row.get("state", "").strip() or None,
                postal_code=row.get("postal_code", "").strip() or None,
                country=row.get("country", "").strip() or "US",
                preferred_contact=row.get("preferred_contact", "").strip() or "email",
                tax_exempt=_parse_bool(row.get("tax_exempt", "")),
                notes=row.get("notes", "").strip() or None,
            )
            db.session.add(customer)
            imported += 1
        except Exception as e:
            errors.append({"row": i, "message": str(e)})

    if imported > 0:
        db.session.commit()

    return {"imported": imported, "skipped": skipped, "errors": errors}


def _import_mapped_inventory(rows):
    """Import inventory rows using mapped field keys."""
    imported = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        try:
            sku = row.get("sku", "").strip()
            if sku:
                existing = InventoryItem.query.filter_by(sku=sku).first()
                if existing:
                    skipped += 1
                    continue

            name = row.get("name", "").strip()
            if not name:
                errors.append({"row": i, "message": "Name is required"})
                continue

            category = row.get("category", "").strip()
            if not category:
                errors.append({"row": i, "message": "Category is required"})
                continue

            item = InventoryItem(
                sku=sku or None,
                name=name,
                category=category,
                subcategory=row.get("subcategory", "").strip() or None,
                manufacturer=row.get("manufacturer", "").strip() or None,
                purchase_cost=_parse_decimal(row.get("purchase_cost", "")),
                resale_price=_parse_decimal(row.get("resale_price", "")),
                quantity_in_stock=_parse_decimal(row.get("quantity", "")) or Decimal("0"),
                reorder_level=_parse_decimal(row.get("reorder_level", "")) or Decimal("0"),
                unit_of_measure=row.get("unit", "").strip() or "each",
                storage_location=row.get("location", "").strip() or None,
                notes=row.get("notes", "").strip() or None,
            )
            db.session.add(item)
            imported += 1
        except Exception as e:
            errors.append({"row": i, "message": str(e)})

    if imported > 0:
        db.session.commit()

    return {"imported": imported, "skipped": skipped, "errors": errors}


# ---------------------------------------------------------------------------
# Legacy fixed-column import (backward compatible)
# ---------------------------------------------------------------------------

def parse_csv(file_content, entity_type):
    """Parse a CSV string and return structured rows with validation.

    Args:
        file_content: CSV string content (may include BOM).
        entity_type: "customers" or "inventory".

    Returns:
        dict with keys:
          - "headers": list of detected column headers
          - "rows": list of dicts (one per data row)
          - "errors": list of {"row": N, "message": str}
          - "preview": first 10 rows as dicts
    """
    # Strip BOM if present
    if file_content.startswith("\ufeff"):
        file_content = file_content[1:]

    reader = csv.DictReader(io.StringIO(file_content))
    if not reader.fieldnames:
        return {"headers": [], "rows": [], "errors": [{"row": 0, "message": "Empty or invalid CSV file"}], "preview": []}

    headers = list(reader.fieldnames)
    rows = []
    errors = []

    for i, row in enumerate(reader, start=2):  # Row 2 = first data row
        rows.append(row)

    # Validate
    if entity_type == "customers":
        errors = _validate_customer_rows(rows)
    elif entity_type == "inventory":
        errors = _validate_inventory_rows(rows)

    return {
        "headers": headers,
        "rows": rows,
        "errors": errors,
        "preview": rows[:10],
    }


def import_customers(rows):
    """Import validated customer rows into the database.

    Args:
        rows: list of dicts from parse_csv.

    Returns:
        dict: {"imported": N, "skipped": N, "errors": [...]}
    """
    imported = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        try:
            # Skip rows with existing email (if email provided)
            email = _get(row, "Email", "email")
            if email:
                existing = Customer.query.filter_by(email=email).first()
                if existing:
                    skipped += 1
                    continue

            customer = Customer(
                customer_type=_get(row, "Type", "type") or "individual",
                first_name=_get(row, "First Name", "first_name") or "",
                last_name=_get(row, "Last Name", "last_name") or "",
                business_name=_get(row, "Business Name", "business_name"),
                contact_person=_get(row, "Contact Person", "contact_person"),
                email=email,
                phone_primary=_get(row, "Phone", "phone"),
                address_line1=_get(row, "Address", "address"),
                city=_get(row, "City", "city"),
                state_province=_get(row, "State", "state"),
                postal_code=_get(row, "Postal Code", "postal_code"),
                country=_get(row, "Country", "country") or "US",
                preferred_contact=_get(row, "Preferred Contact", "preferred_contact") or "email",
                tax_exempt=_parse_bool(_get(row, "Tax Exempt", "tax_exempt")),
                notes=_get(row, "Notes", "notes"),
            )
            db.session.add(customer)
            imported += 1
        except Exception as e:
            errors.append({"row": i, "message": str(e)})

    if imported > 0:
        db.session.commit()

    return {"imported": imported, "skipped": skipped, "errors": errors}


def import_inventory(rows):
    """Import validated inventory rows into the database.

    Args:
        rows: list of dicts from parse_csv.

    Returns:
        dict: {"imported": N, "skipped": N, "errors": [...]}
    """
    imported = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        try:
            sku = _get(row, "SKU", "sku")
            if sku:
                existing = InventoryItem.query.filter_by(sku=sku).first()
                if existing:
                    skipped += 1
                    continue

            name = _get(row, "Name", "name")
            if not name:
                errors.append({"row": i, "message": "Name is required"})
                continue

            category = _get(row, "Category", "category")
            if not category:
                errors.append({"row": i, "message": "Category is required"})
                continue

            item = InventoryItem(
                sku=sku or None,  # None for empty to avoid unique constraint on ''
                name=name,
                category=category,
                subcategory=_get(row, "Subcategory", "subcategory"),
                manufacturer=_get(row, "Manufacturer", "manufacturer"),
                purchase_cost=_parse_decimal(_get(row, "Purchase Cost", "purchase_cost")),
                resale_price=_parse_decimal(_get(row, "Resale Price", "resale_price")),
                quantity_in_stock=_parse_decimal(_get(row, "Quantity", "quantity")) or Decimal("0"),
                reorder_level=_parse_decimal(_get(row, "Reorder Level", "reorder_level")) or Decimal("0"),
                unit_of_measure=_get(row, "Unit", "unit") or "each",
                storage_location=_get(row, "Location", "location"),
                notes=_get(row, "Notes", "notes"),
            )
            db.session.add(item)
            imported += 1
        except Exception as e:
            errors.append({"row": i, "message": str(e)})

    if imported > 0:
        db.session.commit()

    return {"imported": imported, "skipped": skipped, "errors": errors}


# ---------------------------------------------------------------------------
# Validation helpers
# ---------------------------------------------------------------------------

def _validate_customer_rows(rows):
    """Validate customer rows and return a list of errors."""
    errors = []
    for i, row in enumerate(rows, start=2):
        ctype = _get(row, "Type", "type") or "individual"
        if ctype == "individual":
            if not _get(row, "First Name", "first_name") and not _get(row, "Last Name", "last_name"):
                errors.append({"row": i, "message": "Individual customers require first or last name"})
        elif ctype == "business":
            if not _get(row, "Business Name", "business_name"):
                errors.append({"row": i, "message": "Business customers require a business name"})
    return errors


def _validate_inventory_rows(rows):
    """Validate inventory rows and return a list of errors."""
    errors = []
    for i, row in enumerate(rows, start=2):
        if not _get(row, "Name", "name"):
            errors.append({"row": i, "message": "Name is required"})
        if not _get(row, "Category", "category"):
            errors.append({"row": i, "message": "Category is required"})
        cost = _get(row, "Purchase Cost", "purchase_cost")
        if cost:
            try:
                Decimal(cost)
            except InvalidOperation:
                errors.append({"row": i, "message": f"Invalid purchase cost: {cost}"})
    return errors


# ---------------------------------------------------------------------------
# Value parsing helpers
# ---------------------------------------------------------------------------

def _get(row, *keys):
    """Get a value from a row dict, trying multiple key names."""
    for key in keys:
        val = row.get(key, "").strip() if row.get(key) else ""
        if val:
            return val
    return ""


def _parse_bool(val):
    """Parse a boolean value from string."""
    if not val:
        return False
    return val.lower() in ("true", "yes", "1", "y")


def _parse_decimal(val):
    """Parse a decimal value from string, returning None for empty/invalid."""
    if not val:
        return None
    val = val.strip() if isinstance(val, str) else val
    if not val:
        return None
    try:
        return Decimal(val)
    except InvalidOperation:
        return None
